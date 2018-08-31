import tkinter as tk
import pandas as pd
import openpyxl as pyx
import re
import csv
import os
import shutil
import time
from tkinter import messagebox as tkMessageBox
from tkinter import filedialog as tkFileBox
from tkinter import ttk as tttk



class Application(tk.Frame):

    def __init__(self, master):
        self.master = master
        super().__init__(self.master)
        self.master.title("CSV Application")
        self.master.geometry("300x138")
        self.create_widgets()

    def create_widgets(self):
        self.hi_there = tk.Radiobutton(self)
        self.hi_there["text"] = "Please Choose CSV File"
        self.hi_there["indicatoron"] = 0
        self.hi_there["command"] = self.button_press         #the brackets mean its a command or option
        self.hi_there["width"]= 300
        self.hi_there["font"]= ("Times New Roman", 15)
        self.hi_there.pack(side="top")
        self.radio = tk.Radiobutton(self)
        self.radio["text"] = "View Contents"
        self.radio["indicatoron"] = 0
        self.radio["width"] = 300
        self.radio["font"]= ("Times New Roman", 15)
        self.radio["command"]=self.read_new
        self.radio.pack(side="top")
        self.add=tk.Radiobutton(self, text="Add To CSV", indicatoron=0, width=300, command=self.add_new, font=("Times New Roman", 15))
        self.add.pack(side="top")
        self.holder=tk.Radiobutton(self, text="Export To Excel", indicatoron=0, width=300, font=("Times New Roman", 15), fg="red", command=self.export_to_excel)
        self.holder.pack(side="bottom")
        self.pack()

    def button_press(self):
        self.csv_file = tkFileBox.askopenfile(title="Please select a csv file", filetypes=[("CSV Files", "*.csv")])
        while True:
            if self.csv_file == None:
                self.csv_file = tkFileBox.askopenfile(title="Please chose a CSV file", filetypes=[("CSV Files", "*.csv")])
            else:
                break
        self.a = open_csv(self.csv_file)       #by reading the csv file here the variable gets stored in memory for the class


    #this reads in and tests for the new file and if it works it runs the display_new window
    def read_new(self):
        try:
           self.a
        except AttributeError:
            tkMessageBox.showerror("Error", "You have not selected a CSV file") 
        else:
            self.display_new()

    def display_new(self):
        y=0
        x=0
        list_b= self.a.display_excel()             #sets the list to be the return list of class function test

        window = tk.Toplevel()
        window.geometry("400x500")
        window.title("Displayed CSV")


        sc = tk.Scrollbar(window)                    #create a scrollbar on window
        hc= tk.Scrollbar(window, orient="horizontal")
        sc.pack(side= "right", fill="y")
        hc.pack(side="bottom", fill="x")             
        canvas=tk.Canvas(window, borderwidth=0, yscrollcommand=sc.set, xscrollcommand=hc.set)   #creates a canvas on window with a yscrollcommand
        canvas.pack(side="left", expand=True, fill="both")
        sc.config(command=canvas.yview)                        #config the scrollbar to have a y view config    
        hc.config(command=canvas.xview)
        
        Frame = tk.Frame(canvas, background="white")                                 #create a frame to put into canvas
        canvas.create_window((0,0), window=Frame,anchor="nw")    #put the frame inside the canvas with create window  
    

        for i in self.a.display_columns():             # This creates labels for all the column names on row 0
            e1=tk.Label(Frame, bg="white",width=10, borderwidth=2, relief="solid", text=i)
            e1.grid(row=0, column= y)
            y+=1

        for i in range(len(list_b)):      # This creates all the data from the actual excel file. Creates a grid i column names wide with j rows for data in the actual arrays stored
            x=len(list_b[i]) 
            for j in range(len(list_b[i])):
                e2=tk.Label(Frame, bg="white",width=10, borderwidth=1, relief="ridge", text=list_b[i][j]) #takes the list gets the array and then the values in the array
                e2.grid(row=j+1,column=i)

        canvas.config(scrollregion=(0,0,y*80,x*20))      #The grid is 400/500 so you have to adjust how much displays per via scrolllregion


    def add_new(self):         #tests to see if file exists for a adding new button
        try:
           self.a
        except AttributeError:
            tkMessageBox.showerror("Error", "You have not selected a CSV file")
        else:
            self.data_add()

    def data_add(self):                  #actual data and processing for the add_new button
        list_new = self.a.display_columns()
        self.entries=[]                  #this is going to create a list to hold all of the entries
        window = tk.Toplevel()
        window.title("Adding Data")
        window.geometry("250x800")       #create a frame for everything to store, it will look better, frame for combo box, and for everything else

        frame_1 = tk.Frame(window, width=500, height=700)
        frame_1.pack(side="top")

        frame_3 = tk.Frame(window, width=300, height=100)
        frame_3.pack(side="bottom", pady=10)

        for i in range(len(list_new)):      #This creates the entrys
            test = tk.Label(frame_1, text=list_new[i], background="white", font=("Times New Roman", 15),width=10, borderwidth=2, relief="solid")
            test.pack(side="top", pady=20)
            self.entries.append(tk.Entry(frame_1, background="white"))   #This creates the number of entrys per range and adds it to the overall list of entries
            self.entries[i].pack(side="top")
 
        Button = tk.Button(frame_3, text="Press To Add", width=15, background="#42dcf4", command=self.add_button, font=("Times New Roman", 15))
        Button.pack(side="bottom", ipady=(5))

        Button_1 = tk.Button(frame_3, text="Import list", width=15, background="#f5fc92", command=self.add_list, font=("Times New Roman", 15), fg="red")
        Button_1.pack(side="top", ipady=(5), pady=(20))

    def add_button(self):
        list_new = self.a.display_columns()
        list_values = []

        for i in range(len(list_new)):
            list_values.append(self.entries[i].get())     #This appends all the entry data to a list
            self.entries[i].delete(0, "end")
        try:
            with open(self.csv_file.name, "a") as f:      #this opens the file and appends the row to it
                writer = csv.writer(f)
                writer.writerow(list_values)
        except AttributeError:
            print("Something went wrong")
        else:
            self.a
            tkMessageBox.showinfo("Success!", "Your data has been entered")

    def add_list(self):    ##work in progress
        new_csv= tkFileBox.askopenfile(title="Please select a csv file", filetypes=[("CSV Files", "*.csv")])
        csv_file = open_csv(new_csv)
        print(csv_file.display_columns())

    def export_to_excel(self):
        try:
           self.a
        except AttributeError:
            tkMessageBox.showerror("Error", "You have not selected a CSV file")
        else:
            self.export_to()
            

    def export_to(self):

        self.excel_file = tkFileBox.askopenfile(title="Please select a excel template", filetypes=[("Excel Files", "*.xlsx")])
  
        if self.excel_file != None:
            self.new_window=tk.Toplevel()
            self.new_window.title("Type file name")
            self.new_window.geometry("300x200")
            test = tk.Label(self.new_window, text="Save File Name", background="white", font=("Times New Roman", 15), width=15, borderwidth=2, relief="solid")
            test.pack(side="top")
            self.new_entry= tk.Entry(self.new_window, background="white", font=("Times New Roman", 12))
            self.new_entry.pack(side="top", pady=30)
            Button = tk.Button(self.new_window, text="Press To Export", width=15, background="#42dcf4", command=self.export_to_command, font=("Times New Roman", 15))
            Button.pack(side="top", pady=10)
    
    def export_to_command(self):
        self.new_window.withdraw()
        self.master.withdraw()

        frame_new = tk.Toplevel()                                      #creates new frame for the progressbar
        frame_new.geometry("250x25")
        frame_new.title("Export Loading")
        progress = tttk.Progressbar(frame_new, orient=tk.HORIZONTAL, mode="indeterminate", length=250)
        progress.pack(side="top")

        test = self.excel_file.name
        file_name = self.new_entry.get() + ".xlsx"
        b = to_Excel(test)
        b.new_excel(file_name)             #Runs the commands to generate a new file based on name user chose

        values = b.get_all_names()
        names = self.a.display_columns()
        for i in values:
            progress.step(5)
            for x_keys in i.keys():                      # parses the dictionaries in the lists
                if x_keys in names:
                    row_test = i[x_keys]
                    row_number = int(row_test[:-2])         #parses the row key for an integer
                    col_test = i[x_keys]
                    col_number = int(col_test[-2:]) + 1    #parses the row key for the column and adds 1 because of excel 
                    wb = pyx.load_workbook(file_name)
                    sheet = wb.active
                    for j in self.a.get_values(x_keys):
                        frame_new.update()                  #By updating the progress bar here it continously updates the frame
                        row_number += 1
                        sheet.cell(row=row_number, column=col_number).value = j
                    wb.save(file_name)                     #doing wb.save forces the file to be saved after its done

        frame_new.destroy()
        new_message = tkMessageBox.askquestion("Finished!", "The data has been exported", type="ok")
        
        if new_message == "ok":
            self.master.deiconify()
            self.new_window.deiconify()

        #After this you can do a new tkinter window that says "Your export is done"
      
        
class open_csv():
    def __init__(self,csv):
        self.csv = csv
        self.everything = pd.read_csv(self.csv)
        
    
    def display_columns(self):
        list_new= []
        for test in list(self.everything.columns.values):
            list_new.append("%s"%test) 
        return list_new                                 ##displays columns in the csv


    def display_excel(self):    #creates a list using the columns it looks up the values in the columns and seperates them into seperate arrays
        list_test=[]
        for i in self.everything.columns:
            list_test.append(list(self.everything.loc[:, i].values))
        return list_test

    def get_values(self, search_variable):
      list_value =  self.everything.loc[:, search_variable].values  
      return list_value  

class to_Excel():
    def __init__(self, excel_name):
        self.excel_name= excel_name
        self.actual_excel = pd.read_excel(excel_name)


    def get_all_names(self):
        names = []
        wb = pyx.load_workbook(self.excel_name)
        sheet = wb.active
        for i in range(1, sheet.max_row + 1):
            for j in range(sheet.max_column):
                if sheet[i][j].value:
                    s = str(i) + "   " + str(j)
                    names.append({sheet[i][j].value: s})      ## Creates a dictionary of the name and its actual location
        return names


    def display_columns(self):
        list_new = []
        for i in self.actual_excel.columns.values:       
            list_new.append(i)
        return list_new                                         ##returns list that shows all the column names

    def new_excel(self, excel_file):                         #This creates a new excel file copy of the file chosen
        shutil.copy(self.excel_name, excel_file)

def main():
    root = tk.Tk()
    app = Application(root)
    app.mainloop()

#    a=open_csv("Test.csv")
#    print(a.get_values("Hello"))

#    a=to_Excel("Book1.xlsx")
#    print(a.get_all_names())



if __name__ == "__main__":
    main()